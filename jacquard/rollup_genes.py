import argparse
import numpy
import os
from os import listdir
from os.path import isfile, join
import pandas as pd
from pandas import *
import random
import re
import sys 
import openpyxl
from openpyxl import load_workbook

SNPEFF_IMPACT_COLUMN = "SNPEFF_TOP_EFFECT_IMPACT"
DBNSFP_IMPACT_DAMAGING_COLUMN = "dbNSFP_Impact_Damaging"

class RollupError(Exception):
    """Base class for exceptions in this module."""
    pass 

def create_df(input, delim):
    initial_df = pd.read_csv(input, sep=delim, header=False, dtype='str')
    
    return initial_df
  
def verify_headers(df):
    epee_missing_columns = []
    epee_cols = ["CHROM", "POS", "REF", "ANNOTATED_ALLELE", "GENE_SYMBOL"]
    for col in epee_cols:
        if col not in df.columns:
            epee_missing_columns.append(col)
    
    vcf_missing_columns = []
    vcf_cols = ["CHROM", "POS", "REF", "ALT"]
    if epee_missing_columns != []:
        for col in vcf_cols:
            if col not in df.columns:
                vcf_missing_columns.append(col)
    else:
        return epee_cols
        
    if vcf_missing_columns != []:
        raise RollupError("Unable to determine file source. Input file is missing required columns - needs to include either [{0}] for Epee or [{1}] for VCF".format(epee_missing_columns, vcf_missing_columns))
        
    else:
        if "GENE_SYMBOL" in df.columns:
            vcf_cols.append("GENE_SYMBOL")
        return vcf_cols
            

def gene_rollup_highest_impact(initial_df, samples, cols):
    sample_cols =  initial_df.filter(regex=re.escape(samples))
    sample_col_array = sample_cols.columns
    col_array = []
    col_array.extend(sample_col_array)

    columns = cols + [SNPEFF_IMPACT_COLUMN]

    required_columns = set(columns +  col_array)
    for col in initial_df.columns:
        if col not in required_columns:
            del initial_df[col]
    try:
        melted_df = pd.melt(initial_df, id_vars=columns, var_name="Sample", value_name="Sample_Data")
    except Exception as e :
        print e
        raise RollupError("Cannot melt dataframe. {0}".format(e))
        
    melted_df = melted_df.fillna(".")

    filtered_df = melted_df[melted_df["Sample_Data"] != "."]

    pivoted_df = pd.pivot_table(filtered_df, index=["GENE_SYMBOL", "Sample"], columns=[SNPEFF_IMPACT_COLUMN], values=["Sample_Data"], aggfunc=np.count_nonzero, fill_value=0)

    pivoted_df = pivoted_df["Sample_Data"]

    lst = ["HIGH", "MODERATE", "LOW", "MODIFIER"]
    for item in lst:
        if item not in pivoted_df.columns:
            pivoted_df[item] = 0
        pivoted_df[item + "_initial_sum"] = pivoted_df[item].map(int)

    pivoted_df["SnpEff_Impact"] = pivoted_df["HIGH"].apply(lambda x: "h" * x) + pivoted_df["MODERATE"].apply(lambda x: "m" * x) + pivoted_df["LOW"].apply(lambda x: "l" * x) + pivoted_df["MODIFIER"].apply(lambda x: "x" * x)
    pivoted_df["Impact_score"] = pivoted_df["HIGH"] * 100000.0 + pivoted_df["MODERATE"] + pivoted_df["LOW"]/100000.0 + pivoted_df["MODIFIER"]/10**12

    del pivoted_df["HIGH"]
    del pivoted_df["MODERATE"]
    del pivoted_df["LOW"]
    del pivoted_df["MODIFIER"]
    
    expanded_df = pivoted_df.unstack()
    
    return expanded_df

def gene_rollup_damaging_impact(initial_df, samples, cols):   
    sample_cols =  initial_df.filter(regex=re.escape(samples))
    sample_col_array = sample_cols.columns
    col_array = []
    col_array.extend(sample_col_array) 
    
    if DBNSFP_IMPACT_DAMAGING_COLUMN in initial_df.columns:
        initial_df.rename(columns={DBNSFP_IMPACT_DAMAGING_COLUMN: "dbNSFP_rollup_damaging"}, inplace=True)
    
    columns = cols + ["dbNSFP_rollup_damaging"]
    
    required_columns = set(columns +  col_array)
    for col in initial_df.columns:
        if col not in required_columns:
            del initial_df[col]

    try:
        melted_df = pd.melt(initial_df, id_vars=columns, var_name="Sample", value_name="Sample_Data")
    except Exception as e :
        raise RollupError("Cannot melt dataframe. {0}".format(e))
    
    filtered_df = melted_df.fillna(0)
    
    filtered_df = filtered_df[filtered_df["dbNSFP_rollup_damaging"] != "0"]
    filtered_df = filtered_df[filtered_df["Sample_Data"] != 0]

    try:
        filtered_df["dbNSFP_rollup_damaging"] = filtered_df["dbNSFP_rollup_damaging"].apply(lambda x: int(x))
        df_impact = pd.pivot_table(filtered_df, index=["GENE_SYMBOL"], columns=["Sample"], values=["dbNSFP_rollup_damaging"], aggfunc=np.sum)
    except:
        df_impact = pd.pivot_table(filtered_df, index=["GENE_SYMBOL"], columns=["Sample"], values=["dbNSFP_rollup_damaging"], aggfunc=lambda x: "*")
    
    return df_impact
    
def combine_dfs(df1, df2):
    df_combined = df1.join(df2, how='outer')
    df_combined = df_combined.fillna(0)
    
    ranked_df = calculate_ranks(df_combined)
    
    ranked_df = ranked_df.fillna("")
    
    return ranked_df
    
def calculate_ranks(df):
    df["SnpEff_Impact_Rank"] = df["Impact_score"].sum(axis=1)
    df["dbNSFP_rollup_damaging_Rank"] = df["dbNSFP_rollup_damaging"].sum(axis=1)
    
    lst = ["HIGH", "MODERATE", "LOW", "MODIFIER"]

    for item in lst:
        df[item + "_sum"] = df[item + "_initial_sum"].sum(axis=1)

    df["dbNSFP_rollup_damaging"] = df["dbNSFP_rollup_damaging"].applymap(lambda x: "" if x == 0.0 else x)
    
    df["Impact_Score"] = df["SnpEff_Impact_Rank"]
    
    df = df.sort("SnpEff_Impact_Rank", ascending=0)
    df = df.sort("dbNSFP_rollup_damaging_Rank", ascending=0)
    
    df["SnpEff_Impact_Rank"] = df["SnpEff_Impact_Rank"].rank(ascending=0, method="min")
    df["dbNSFP_rollup_damaging_Rank"] = df["dbNSFP_rollup_damaging_Rank"].rank(ascending=0, method="min")

    del df["Impact_score"]
    lst = ["HIGH", "MODERATE", "LOW", "MODIFIER"]
    for item in lst:
        del df[item + "_initial_sum"]
    
    return df   
 
def rearrange_columns(output_df):
    ##change tuples to strings:
    lst = []
    for i in list(output_df.columns.values):
        if type(i) is tuple:
            i = "_".join(i)
        lst.append(i)
    output_df.columns = lst 
    
    ##change order of columns:
    lst = change_order(lst)    
    output_df = output_df.ix[:,lst]
    del output_df["Impact_Score_"]
    
    return output_df
    
def change_order(lst):
    gene_lst = []
    sum_lst = []
    level_lst = []
    impact_lst = []
    score_lst = []
    
    for i, header in enumerate(lst):    
        if re.search("GENE_SYMBOL", header):
            gene_lst.append(header)
        elif re.search("SnpEff_Impact", header):
            impact_lst.append(header)
        elif re.search("sum", header):
            sum_lst.append(header)
        elif re.search("Score", header):
            score_lst.append(header)
        elif re.search("dbNSFP_rollup_damaging", header):
            level_lst.append(header)
   
    lst = gene_lst + impact_lst + sum_lst + score_lst + level_lst
    
    return lst

def style_workbook(output):
    wb = load_workbook(output)
    ws = wb.active
    
    for col in ws.columns:
        if re.search("Rank", col[0].value):
            format_cell_range(ws, col, '3366FF', 'FFFFFFFF')
        elif re.search("HIGH", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', 'FF0000')
        elif re.search("MODERATE", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', 'FFCC00')
        elif re.search("LOW", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', '00FF00')
        elif re.search("MODIFIER", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', 'FF6600')

    wb.save(output)
    
def format_cell_range(ws, column, min_color, max_color):
    first_cell = str(column[1]).strip("(<>)").split(".")[-1] 
    last_cell = str(column[-1]).strip("(<>)").split(".")[-1] 
    cell_range = first_cell + ":" + last_cell

    ws.conditional_formatting.add2ColorScale(cell_range, 'min', None, min_color, 'max', None, max_color)
    
def process_files(input, output, samples, delim):
    initial_df = create_df(input, delim)
    
    cols = verify_headers(initial_df)
    
    second_df = create_df(input, delim)
    
    df1 = gene_rollup_highest_impact(initial_df, samples, cols)
    df2 = gene_rollup_damaging_impact(second_df, samples, cols)
    
    rollup_df = combine_dfs(df1, df2)
    
    rearranged_df = rearrange_columns(rollup_df)
    
    new_columns = []
    for col in rearranged_df.columns:
        new_col = re.sub(samples + "_", "", col)
        new_columns.append(new_col)
        
        # if col[-1] == "_":
            # new_columns.append(col[:-1])
    rearranged_df.columns = new_columns
    
    print "writing to excel file: {0}".format(output)
    writer = ExcelWriter(output)
    
    rearranged_df.to_excel(writer, "Variant_output", index=True, merge_cells = 0)  

    print "saving file"
    writer.save() 
    
    # print "styling workbook"
    # style_workbook(output)

    print "done"
    

def add_subparser(subparser):
    parser_rollup = subparser.add_parser("rollup_genes", help="Summarizes variant-level Excel file to create a gene-level rollup. Returns an Excel file containing concatenation of all input files.")
    parser_rollup.add_argument("input_file", help="Path to input variant-level tab-separated file")
    parser_rollup.add_argument("output_file", help="Path to output gene-level XLSX file")
    parser_rollup.add_argument("-s", "--sample_identifier", required=True,
            help="Identifier for all samples. This should be a string that is present in all samples.")
    parser_rollup.add_argument("-d", "--input_delimiter",
            help="Delimiter for input file. If no delimiter given, defaults to comma.")

def execute(args, execution_context):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pd.set_option('chained_assignment', None)
    
    input   = os.path.abspath(args.input_file)
    output  = os.path.abspath(args.output_file)
    samples = args.sample_identifier
    delim   = args.input_delimiter if args.input_delimiter else "\t"
# 
    if not os.path.isfile(input):
        print "Error. Specified input file {0} does not exist".format(input)
        exit(1)
 
    process_files(input, output, samples, delim)

def main():
    parser = argparse.ArgumentParser(
        usage="jacquard",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description='''---''',
        epilog="---")

    subparsers = parser.add_subparsers(title="subcommands",
                                       dest="subparser_name")
    add_subparser(subparsers)
    args = parser.parse_args(sys.argv[1:])
    execute(args, [])

if __name__ == '__main__':
    main()