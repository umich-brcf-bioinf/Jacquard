import argparse
import datetime
import numpy
import os
from os import listdir
from os.path import isfile, join
import pandas as pd
from pandas import *
import random
import re
import sys 
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.style import Color, Fill, Font, Alignment

class PivotError(Exception):
    """Base class for exceptions in this module."""
    pass 

def style_workbook(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active

    for col in ws.columns:
        if col[0].value == "IGV":
            format_links(ws, col, "IGV")
        elif col[0].value == "UCSC": 
            format_links(ws, col, "UCSC")
        elif col[0].value == "dbSNP": 
            format_links(ws, col, "dbSNP")
#         count = 0
#         colors = ["FFFF99", "FFCC00", "FF9900", "FF6600"]
        # for tag in pivot_values:
            # if re.search(tag, col[0].value):
                # fill_cell(col, colors[count])
            # count += 1
        
        if re.search("CHROM|POS|ID|REF|ALT|Mult_Alt|Mult_Gene|ANNOTATED_ALLELE|GENE_SYMBOL|IGV|UCSC|dbSNP|QUAL|FILTER|INFO", col[0].value):
            fill_cell(col, "C0C0C0")   

        if col[0].style.fill.start_color.index == "FFFFFFFF":
            fill_cell(col, "D7E8F0")
        if col[0].value == "Mult_Alt":
            fill_mults(ws, col)
        if col[0].value == "Mult_Gene":
            fill_mults(ws, col)
        if col[0].value == "INFO":
            fill_mult_alts(ws, col)

    wb.save(output_file)
 
def format_links(ws, column, cell_value):
    for pos in column:
        if pos != column[0]:
            desired_cell = str(pos).strip("(<>)").split(".")[-1]
            ws[desired_cell].hyperlink = pos.value
            pos.value = cell_value
            pos.style.font.color.index = Color.BLUE
            pos.style.font.underline = Font. UNDERLINE_SINGLE
            
def fill_cell(column, color):
    column[0].style.alignment.text_rotation = 90
    column[0].style.alignment.vertical = Alignment.VERTICAL_BOTTOM
    column[0].style.fill.fill_type = Fill.FILL_SOLID
    column[0].style.fill.start_color.index = color

def fill_row(row, color):
    row.style.fill.fill_type = Fill.FILL_SOLID
    row.style.fill.start_color.index = color
    
def fill_mults(ws, col):
    for row in col:
        if row.value == "True":
            coordinate = row.address
            row = row.address[1:]
            for item in ws.range("A" + row + ":" + coordinate):
                for cell in item:
                    if cell.value != "None":
                        fill_row(cell, "FFD698")
                        
def fill_mult_alts(ws, col):
    for row in col:
        if row.value == "Mult_Alt":
            coordinate = row.address
            row = row.address[1:]
#             for item in ws.range("A" + row + ":" + coordinate):
            for item in ws.range("A" + row + ":D" + row):
                for cell in item:
                    if cell.value != "None":
                        fill_row(cell, "FFD698")
    
def add_subparser(subparser):
    parser_pivot = subparser.add_parser("style", help="Accepts an XLSX file and outputs a styled XLSX file with conditional formatting")
    parser_pivot.add_argument("input_file", help="Input XLSX file that has been formatted by Jacquard")
    parser_pivot.add_argument("output_file", help="Output XLSX file")
     
def execute(args, execution_context):
    input_file = os.path.abspath(args.input_file)
    output_file = os.path.abspath(args.output_file)

    for file in (input_file, output_file):
        fname, extension = os.path.splitext(file)
        if extension != ".xlsx": 
            print "Error. Specified output {0} must have a .vcf extension".format(file)
            exit(1)
    
    if not os.path.isfile(input_file):
        print "Error. Specified input directory {0} does not exist".format(input_file)
        exit(1)
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
#     pd.set_option('chained_assignment', None)

    style_workbook(input_file, output_file)
