#!/usr/bin/python2.7
import argparse
import numpy
import os
from os import listdir
from os.path import isfile, join
import pandas as pd
from pandas import *
import random
import re
import sys 
import openpyxl
from openpyxl import load_workbook
from openpyxl.style import Color, Fill, Font

class PivotError(Exception):
    """Base class for exceptions in this module."""
    pass 


class VariantPivoter():
    MISSING_REQUIRED_COLUMNS_ERROR="The columns of specified dataframe do not " +\
        "contain required columns {0}; review input data."

    @staticmethod
    def _build_transform_method(rows, columns, pivot_values):
        def transform(df):
            expanded_df = expand_format(df, pivot_values, rows)
            
            return project_prepivot(expanded_df, pivot_values, rows, columns)
        return transform

    def __init__(self, rows, cols, pivot_values, combined_df=pd.DataFrame(), annot_df=pd.DataFrame()):
        self._rows = rows
        self._cols = cols
        self._pivot_values = pivot_values
        self._transform = VariantPivoter._build_transform_method(self._rows, self._cols, self._pivot_values)
        self._combined_df = combined_df
        self._annot_df = annot_df

    def pivot(self):
        pivoted_df = pd.pivot_table(
            self._combined_df, values=self._pivot_values, rows=self._rows, cols=self._cols, 
            aggfunc=lambda x: x)

        try:
            pivoted_df = pivoted_df.applymap(lambda x: int(x))
        except:
            pass

        return pivoted_df
 
    def add_file(self, path, reader, header_index):
        initial_df  = create_initial_df(path, reader, header_index)
        
        if "#CHROM" in initial_df.columns:
            initial_df.rename(columns={"#CHROM": "CHROM"}, inplace=True)
        
        self._annot_df = append_to_annot_df(initial_df, self._annot_df)
        
        unpivoted_df = self.is_compatible(initial_df)

        self._combined_df = merge_samples(unpivoted_df, self._combined_df)
        
    def validate_annotations(self):
        grouped_df = self._annot_df.groupby(self._rows)
        self._annot_df = grouped_df.last()
        self._annot_df.reset_index(inplace=True)
        
        return self._annot_df
        
    def validate_sample_data(self):
        grouped = self._combined_df.groupby(self._rows + ["SAMPLE_NAME"])
        group = grouped.groups
        
        for column in self._combined_df:
            for key, val in group.items():
                self.find_non_unique_columns(grouped, column, key, val)
            self.find_non_unique_rows(column)
        return self._combined_df
        
    def find_non_unique_columns(self, grouped, column, key, val):
        if len(val) != 1:
            col_data = []
            for index in val:
                data = grouped.get_group(key).ix[index, column]
                if data not in col_data:
                    col_data.append(data)
            if len(col_data) != 1:
                for index in val:
                    self._combined_df.ix[index, column] = "^"    
                    
    def find_non_unique_rows(self, column):
        count = 0
        for data in self._combined_df[column]:
            if type(data) == np.ndarray:
                self._combined_df.ix[count, column] = "^"
            count += 1
                
    def join_dataframes(self, pivot_df):
        output = merge(self._annot_df, pivot_df, left_on=self._rows, right_index=True, how='outer', sort=False)
        output = output.fillna("")

        return output
    
    def sort_rows(self, df):
        df.reset_index(inplace=True)
        sorted_df = df.sort(self._rows)

        return sorted_df
    
    def insert_mult_alt_and_gene(self, df):
        # if "SnpEff_WARNING/ERROR" in df.columns.values and "SnpEff_WARNING/ERROR" not in self._rows:
            # df.rename(columns={"SnpEff_WARNING/ERROR": "WARNING/ERROR"}, inplace=True)
        if "WARNING/ERROR" in df.columns.values and "WARNING/ERROR" not in self._rows:
            df.rename(columns={"WARNING/ERROR": "SnpEff_WARNING/ERROR"}, inplace=True)
        
        if "SnpEff_WARNING/ERROR" in df.columns.values and "SnpEff_WARNING/ERROR" not in self._rows:
            cols_to_group = self._rows + ["SnpEff_WARNING/ERROR"]
        else:
            cols_to_group = self._rows

        grouped = df.groupby(cols_to_group)
        
        df.set_index(cols_to_group, inplace=True)
        
        if "GENE_SYMBOL" in cols_to_group:
            df.insert(0, "Mult_Gene", "")
            df = self.label_mult(grouped, df, "gene")
            
        df.insert(0, "Mult_Alt", "")
        df = self.label_mult(grouped, df, "alt")
        
        df.reset_index(inplace=True)
        
        return df
     
    def label_mult(self, grouped, df, type):
        preliminary_dict = {}
        mult_dict     = {}
        val_index = ""

        count = 0
        for field in grouped.keys:
            if type == "alt":
                if field == "ANNOTATED_ALLELE" or field == "ALT":
                    val_index = count
            elif type == "gene":
                if field == "GENE_SYMBOL":
                    val_index = count
                    
            count += 1
            
        if val_index == "":
            raise PivotError("Cannot calculate mult-alts and/or mult-genes with given keys {0}.".format(grouped.keys))
            
        mult_dict = self._create_mult_dict(grouped, val_index, preliminary_dict, mult_dict)

        for key, vals in mult_dict.iteritems():
            row_key = key.split("_")

            for val in vals:
                complete_key = self._determine_row_key(row_key, val, val_index)

                if type == "alt":
                    df.loc[complete_key, "Mult_Alt"] = "True"
                elif type == "gene":
                    df.loc[complete_key, "Mult_Gene"] = "True"
        
        return df
   
    def _create_mult_dict(self, grouped, val_index, preliminary_dict, mult_dict):
        for k, gp in grouped:
            k_list = []
            for item in k:
                k_list.append(item)
            
            val = str(k_list[val_index])
            del k_list[val_index]
            key = "_".join(k_list)

            if key in preliminary_dict:
                if key in mult_dict:
                    mult_dict[key].append(val)
                else:
                    mult_dict[key] = [preliminary_dict[key], val]
            else:
                preliminary_dict[key] = val
                
        return mult_dict
    
    def _determine_row_key(self, row_key, val, val_index):
        complete_key = []
        first = row_key[:val_index]
        last = row_key[val_index:]
        
        for field in first:
            complete_key.append(field)
            
        complete_key.append(val)
        
        for field in last:
            complete_key.append(field)

        complete_key = tuple(complete_key)
       
        return complete_key
                
    def is_compatible(self, initial_df):   
        unpivoted_df = self._transform(initial_df)

        self._check_required_columns_present(unpivoted_df)
        self._check_pivot_is_unique(unpivoted_df)  
       
        return unpivoted_df

    def _check_required_columns_present(self, dataframe):
        required_columns = set(self._rows + self._cols)
        if not required_columns.issubset(dataframe.columns.values):
            print "columns absent"
            raise PivotError("Missing required columns; contact sysadmin.")
         
    def _check_pivot_is_unique(self, dataframe):   
        group = self._rows + self._cols
        grouped_df = dataframe.groupby(group)

        if len(grouped_df.groups) != len(dataframe):
            print "not unique"
            raise PivotError("Duplicate keys would result in an invalid pivot; contact sysadmin.")

def validate_parameters(input_keys, first_line, header_names, pivot_values):
    invalid_fields = []

    fields = header_names.split("\t")

    for key in input_keys:
        if key not in fields:
            if key == "SnpEff_WARNING/ERROR":
                input_keys.remove(key)
                input_keys.append("WARNING/ERROR")
            else:
                invalid_fields.append(key)
    
    invalid_tags = validate_format_tags(first_line, pivot_values, fields)
                
    message = "Invalid input parameter(s) "
    raise_err = 0
    if invalid_fields != []:
        message += str(invalid_fields)
        raise_err = 1
    if invalid_tags != []:
        message += str(invalid_tags)
        raise_err = 1
    
    return raise_err, message

def validate_format_tags(first_line, pivot_values, fields):
    invalid_tags = []
    for line in first_line:
        my_line = line.split("\t")
        for index, field in enumerate(fields):
            if field == "FORMAT":
                format = my_line[index]
                format_tags = format.split(":")
                
                for val in pivot_values:
                    if val not in format_tags:
                        invalid_tags.append(val)
    
    return invalid_tags
    
def build_pivoter(path, reader, input_keys, pivot_values, header_index):
    initial_df  = create_initial_df(path, reader, header_index)

    if "SnpEff_WARNING/ERROR" in initial_df.columns.values or "WARNING/ERROR" in initial_df.columns.values:
        initial_df.rename(columns={"WARNING/ERROR": "SnpEff_WARNING/ERROR"}, inplace=True)
        initial_df = _exclude_errors_and_warnings(initial_df)
    
    pivoter = VariantPivoter(input_keys, ["SAMPLE_NAME"], pivot_values)
    pivoter.is_compatible(initial_df)
    
    return pivoter
  
def _exclude_errors_and_warnings(df):
        try:
            filtered_df = df[df['SnpEff_WARNING/ERROR']=='.']
            return filtered_df
        except KeyError as e:
            raise PivotError("File is missing SnpEff_WARNING/ERROR column; review input files")
        else:
            raise

def create_initial_df(path, reader, header_index):
    initial_df = pd.read_csv(reader, sep="\t", header=header_index, dtype='str')

    return initial_df
  
def append_to_annot_df(initial_df, annot_df):  
    format_index = initial_df.columns.get_loc("FORMAT")
    last_column = initial_df.shape[1]
    
    annotation_columns = []
    annotation_columns.extend(initial_df.columns)
    
    ##remove column if it is format column or any column to the right of format
    for i in range(format_index, last_column):
        annotation_columns.remove(initial_df.columns[i])
    if "INFO" in annotation_columns:
        annotation_columns.remove("INFO")
    
    temp_df = pd.DataFrame()
    for col in initial_df.columns:
        temp_df[col] = initial_df[col]
        if col not in annotation_columns:
            del temp_df[col]
            
    if annot_df.empty:
        annot_df = temp_df
    else:
        annot_df = annot_df.append(temp_df, ignore_index=True)
    return annot_df
  
def melt_samples(df): 
    format_data = df.ix[0, "FORMAT"].split(":")
    
    first_sample_column_index = df.columns.get_loc("FORMAT") + 1
    first_sample_column = df.ix[:, first_sample_column_index]
    last_column = df.shape[1]
    
    sample_names = []
    for i in range(first_sample_column_index, last_column):
        field_name = df.columns[i]
        data = df.ix[0, field_name].split(":")
        if len(data) == len(format_data):
            sample_names.append(field_name)

    column_list = []
    column_list.extend(df.columns)

    for col in sample_names:
        if col in column_list: 
            column_list.remove(col)   

    melted_df = melt(df, id_vars=column_list, var_name='SAMPLE_NAME', value_name='SAMPLE_DATA')
    
    return melted_df
    
def expand_format(df, formats_to_expand, rows):     
    df = melt_samples(df)

    df["aggregate_format_sample"] = df["FORMAT"] + "=" + df['SAMPLE_DATA']
    df["aggregate_format_sample"] = df["aggregate_format_sample"].map(combine_format_values)
   
    s = df["aggregate_format_sample"].apply(pd.Series, 1).stack()
    s.index = s.index.droplevel(-1)
    s.name = "format_sample"

    unpivoted_format_value_df = s.apply(lambda x: pd.Series(x))
    unpivoted_format_value_df.columns = ["FORMAT2", "VALUE2"]
    
    joined_df = df.join(unpivoted_format_value_df)
    del joined_df["aggregate_format_sample"]
    
    if "#CHROM" in joined_df.columns:
        joined_df.rename(columns={"#CHROM": "CHROM"}, inplace=True)
        
    if "SnpEff_WARNING/ERROR" in joined_df.columns and "SnpEff_WARNING/ERROR" not in rows:
        joined_df.rename(columns={"SnpEff_WARNING/ERROR": "WARNING/ERROR"}, inplace=True)

    try:
        pivoted_df = pd.pivot_table(joined_df, rows=rows+["SAMPLE_NAME"], cols="FORMAT2", values="VALUE2", aggfunc=lambda x: x)
    except:
        raise PivotError("Cannot pivot data.")

    for tag in formats_to_expand:
        if tag not in pivoted_df.columns:
            raise PivotError("Error: specified format tag {0} was not found in the file\n".format(tag))

    pivoted_df.reset_index(inplace=True)

    return pivoted_df

def combine_format_values(aggregate_col):
    pairs = aggregate_col.split("=")
    return zip(pairs[0].split(":"), pairs[1].split(":"))
    
def project_prepivot(df, pivot_values, rows, columns):
    required_columns = set(pivot_values + rows + columns)

    for col in df.columns:
        if col not in required_columns:
            del df[col]

    return df

def merge_samples(reduced_df, combined_df):
    if combined_df.empty:
        combined_df = reduced_df
    else:
        if len(combined_df.columns) == len(reduced_df.columns):
            combined_df = combined_df.append(reduced_df, ignore_index=True)
        else:
            raise PivotError("Columns do not match across files")

    return combined_df    
    
def rearrange_columns(output_df, pivot_values):
    ##change tuples to strings:
    lst = []
    for i in list(output_df.columns.values):
        if type(i) is tuple:
            i = "_".join(i)
        lst.append(i)
    output_df.columns = lst 
    
    ##change order of columns:
    lst = change_order(lst, pivot_values)    
    output_df = output_df.ix[:,lst]

    return output_df
    
def change_order(lst, pivot_values):
    all_pivot_values_lst = []
    for pivot_value in pivot_values:
        pivot_value_lst = []
        for i, header in enumerate(lst):    
            if re.search(pivot_value, header):
                pivot_value_lst.append(header)
        
        all_pivot_values_lst.extend(pivot_value_lst)
    
    meta_lst = []
    annot_lst = []
    for i, header in enumerate(lst):    
        if re.search("CHROM|POS|ID|REF|ALT|ANNOTATED_ALLELE|GENE_SYMBOL|Mult_Alt|Mult_Gene|IGV|UCSC", header):
            meta_lst.append(header)
        elif header not in all_pivot_values_lst:
            annot_lst.append(header) 
                     
    lst = meta_lst + all_pivot_values_lst + annot_lst

    return lst
    
def insert_links(joined_output):
    for row, col in joined_output.T.iteritems():
        joined_output.loc[row, "UCSC"] = '=hyperlink("http://genome.ucsc.edu/cgi-bin/hgTracks?db=hg19&position=chr' + joined_output.loc[row, "CHROM"] + ":" + str(int(joined_output.loc[row, "POS"]) - 50)+ "-" + str(int(joined_output.loc[row, "POS"]) + 50) + '", "UCSC")'
        
        joined_output.loc[row, "IGV"] = '=hyperlink("localhost:60151/goto?locus=chr' + joined_output.loc[row, "CHROM"] + ':' + joined_output.loc[row, "POS"] + '", "IGV")'
 
def style_workbook(output_path):
    wb = load_workbook(output_path)
    ws = wb.active
    
    for col in ws.columns:  
        if re.search("IGV", col[0].value):
            format_links(ws, col, "IGV")
        elif re.search("UCSC", col[0].value): 
            format_links(ws, col, "UCSC")
    
        count = 0
        colors = ["FFFF99", "FFCC00", "FF9900", "FF6600"]
        for tag in pivot_values:
            if re.search(tag, col[0].value):
                fill_cell(col, colors[count])
            count += 1
        
        if re.search("CHROM|POS|ID|REF|ALT|Mult_Alt|Mult_Gene|ANNOTATED_ALLELE|GENE_SYMBOL|IGV|UCSC", col[0].value):
            fill_cell(col, "C0C0C0")   

        if col[0].style.fill.start_color.index == "FFFFFFFF":
            fill_cell(col, "C9F2C9")
        if re.search("Mult_Alt", col[0].value):
            fill_mults(ws, col)
        if re.search("Mult_Gene", col[0].value):
            fill_mults(ws, col)

    wb.save(output_path)
 
def format_links(ws, column, cell_value):
    for pos in column:
        if pos != column[0]:
            desired_cell = str(pos).strip("(<>)").split(".")[-1]
            ws[desired_cell].hyperlink = pos.value
            pos.value = cell_value
            pos.style.font.color.index = Color.BLUE
            pos.style.font.underline = Font. UNDERLINE_SINGLE
            
def fill_cell(column, color):
    column[0].style.fill.fill_type = Fill.FILL_SOLID
    column[0].style.fill.start_color.index = color

def fill_row(row, color):
    row.style.fill.fill_type = Fill.FILL_SOLID
    row.style.fill.start_color.index = color
    
def fill_mults(ws, col):
    for row in col:
        if row.value == "True":
            coordinate = row.address
            row = row.address[1:]
            for item in ws.range("A" + row + ":" + coordinate):
                for cell in item:
                    if cell.value != "None":
                        fill_row(cell, "FFD698")
    
def process_files(sample_file_readers, pivot_builder=build_pivoter):
    first_file_reader = sample_file_readers[0]
    first_path        = str(first_file_reader)
    first_reader      = first_file_reader
    
    print "validating command line parameters"
    raise_err, message = validate_parameters(input_keys, first_line, header_names, pivot_values)  
    if raise_err == 1:
        raise PivotError(message)
        
    print "determining file type"
    pivoter  = pivot_builder(first_path, first_reader, input_keys, pivot_values, header_index)
    
    for file_reader in sample_file_readers:
        path   = str(file_reader)
        reader = file_reader
        
        print "Reading: " + path
         
        pivoter.add_file(path, reader, header_index)
  
    print "validating annotations"
    pivoter.validate_annotations()
    
    print "validating sample data"
    pivoter.validate_sample_data()
  
    print "pivoting data"
    pivoted_df = pivoter.pivot()
    pivoted_df = pivoted_df.fillna("")
    
    print "joining sample data with annotations"
    joined_output = pivoter.join_dataframes(pivoted_df)
    insert_links(joined_output)
  
    print "calculating mult alts and mult genes"
    mult_alt_gene_df = pivoter.insert_mult_alt_and_gene(joined_output)
    
    try:
        mult_alt_gene_df["CHROM"] = mult_alt_gene_df["CHROM"].apply(lambda x: int(x))
        mult_alt_gene_df["POS"] = mult_alt_gene_df["POS"].apply(lambda x: int(x))
    except:
        pass

    print "rearranging columns"
    complete_output = rearrange_columns(mult_alt_gene_df, pivot_values)
    
    print "sorting rows"
    sorted_df = pivoter.sort_rows(complete_output)

    if "index" in sorted_df:
        del sorted_df["index"]
        
    if "WARNING/ERROR" in sorted_df.columns.values:
        sorted_df.rename(columns={"WARNING/ERROR":"SnpEff_WARNING/ERROR"}, inplace=True)

    print "writing to excel file: {0}".format(output_path)
    writer = ExcelWriter(output_path)
    sorted_df.to_excel(writer, "Variant_output", index=False, merge_cells = 0)  
    
    # print "writing to csv file: {0}".format(output_path)
    # sorted_df.to_csv(output_path, index=False, sep=",")  
    # print "almost done..."
   
    print "saving file"
    writer.save() 
    
    print "styling workbook"
    style_workbook(output_path)
   
    print "done"
    
def determine_input_keys(input_dir):
    for file in listdir(input_dir):
        if isfile(join(input_dir, file)):
            fname, extension = os.path.splitext(file)
            
            if extension == ".txt":
                return ["CHROM", "POS", "REF", "ANNOTATED_ALLELE", "GENE_SYMBOL", "SnpEff_WARNING/ERROR"]
           
            elif extension == ".vcf":
                return ["CHROM", "POS", "REF", "ALT"]
                
            else:
                raise PivotError("Cannot determine columns to be used as keys for the pivoting from file [{0}]. Please specify parameter [-k] or [--keys]".format(os.path.abspath(file)))
    
def get_headers_and_readers(input_dir):
    sample_file_readers = []
    headers = []
    header_names = []
    first_line = []
    
    for item in listdir(input_dir):
        if isfile(join(input_dir, item)):
            f = open(input_dir + "/" + item, 'r')
            for num, line in enumerate(f):
                if line.startswith("#"):
                    headers.append(num)
                    header_names.append(line)
                else:
                    first_line.append(line)
                    break
            f.close()
            sample_file_readers.append(input_dir + "/" + item)

    header_index = headers[-1]
    header_names = header_names[-1]
    
    header_names = re.sub(r'#CHROM', 'CHROM', header_names)
    
    return sample_file_readers, header_index, header_names, first_line
    
if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pd.set_option('chained_assignment', None)
    
    parser = argparse.ArgumentParser(
    formatter_class=argparse.RawDescriptionHelpFormatter, 
    description='''\
    Pivot.py
    Pivots input files so that given sample specific information is fielded out into separate columns. Returns an Excel file containing concatenation of all input files. ''', 
    epilog="author: Jessica Bene 05/2014")
    parser.add_argument("input_dir")
    parser.add_argument("output_file")
    parser.add_argument("-k", "--keys",
            help="Columns to be used as keys for the pivoting. Default keys for VCF are CHROM,POS,REF,ALT. Default keys for Epee TXT are CHROM,POS,REF,ANNOTATED_ALLELE,GENE_SYMBOL")
    parser.add_argument("-t", "--tags",
            help="Format tags to be fielded out in the pivoting.")
            
    args   = parser.parse_args()
    input_dir = os.path.abspath(args.input_dir)
    output_path = os.path.abspath(args.output_file)
    input_keys = args.keys.split(",") if args.keys else determine_input_keys(input_dir)
    pivot_values = args.tags.split(",") if args.tags else ["GT"]
    
    output_dir, outfile_name = os.path.split(output_path)
    
    if not os.path.isdir(input_dir):
        print "Error. Specified input directory {0} does not exist".format(input_dir)
        exit(1)
        
    if not os.path.isdir(output_dir):
        os.makedirs(output_dir)
        
    fname, extension = os.path.splitext(outfile_name)
    if extension != ".xlsx": 
        print "Error. Specified output {0} must have a .xlsx extension".format(output_path)
        exit(1)
    
    sample_file_readers, header_index, header_names, first_line = get_headers_and_readers(input_dir)

    process_files(sample_file_readers)
    
    
