#!/usr/bin/python2.7
import argparse
import datetime
import numpy
import os
from os import listdir
from os.path import isfile, join
import pandas as pd
from pandas import *
import random
import re
import sys 
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.style import Color, Fill, Font

class PivotError(Exception):
    """Base class for exceptions in this module."""
    pass 

def style_workbook(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
    
    for col in ws.columns:  
        if re.search("IGV", col[0].value):
            format_links(ws, col, "IGV")
        elif re.search("UCSC", col[0].value): 
            format_links(ws, col, "UCSC")
    
        count = 0
        colors = ["FFFF99", "FFCC00", "FF9900", "FF6600"]
        # for tag in pivot_values:
            # if re.search(tag, col[0].value):
                # fill_cell(col, colors[count])
            # count += 1
        
        if re.search("CHROM|POS|ID|REF|ALT|Mult_Alt|Mult_Gene|ANNOTATED_ALLELE|GENE_SYMBOL|IGV|UCSC", col[0].value):
            fill_cell(col, "C0C0C0")   

        if col[0].style.fill.start_color.index == "FFFFFFFF":
            fill_cell(col, "C9F2C9")
        if re.search("Mult_Alt", col[0].value):
            fill_mults(ws, col)
        if re.search("Mult_Gene", col[0].value):
            fill_mults(ws, col)

    wb.save(output_file)
 
def format_links(ws, column, cell_value):
    for pos in column:
        if pos != column[0]:
            desired_cell = str(pos).strip("(<>)").split(".")[-1]
            ws[desired_cell].hyperlink = pos.value
            pos.value = cell_value
            pos.style.font.color.index = Color.BLUE
            pos.style.font.underline = Font. UNDERLINE_SINGLE
            
def fill_cell(column, color):
    column[0].style.fill.fill_type = Fill.FILL_SOLID
    column[0].style.fill.start_color.index = color

def fill_row(row, color):
    row.style.fill.fill_type = Fill.FILL_SOLID
    row.style.fill.start_color.index = color
    
def fill_mults(ws, col):
    for row in col:
        if row.value == "True":
            coordinate = row.address
            row = row.address[1:]
            for item in ws.range("A" + row + ":" + coordinate):
                for cell in item:
                    if cell.value != "None":
                        fill_row(cell, "FFD698")

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pd.set_option('chained_assignment', None)
    
    parser = argparse.ArgumentParser(
    formatter_class=argparse.RawDescriptionHelpFormatter, 
    description='''\
    Style.py
    Styles a pivoted input XLSX file and outputs a styled XLSX file.''', 
    epilog="author: Jessica Bene 05/2014")
    parser.add_argument("input_file")
    parser.add_argument("output_file")
    
    args   = parser.parse_args()
    input_file = os.path.abspath(args.input_file)
    output_file = os.path.abspath(args.output_file)
    
    if not os.path.isfile(input_file):
        print "Error. Specified input directory {0} does not exist".format(input_file)
        exit(1)
        
    fname, extension = os.path.splitext(input_file)
    if extension != ".xlsx": 
        print "Error. Specified input {0} must have a .xlsx extension".format(input_file)
        exit(1)
    
    fname, extension = os.path.splitext(output_file)
    if extension != ".xlsx": 
        print "Error. Specified output {0} must have an .xlsx extension".format(output_file)
        exit(1)

    print "{0} - styling workbook".format(datetime.fromtimestamp(time.time()).strftime('%Y/%m/%d %H:%M:%S'))
    style_workbook(input_file, output_file)    
    
    
