#!/usr/bin/python2.7
import numpy
import os
from os import listdir
from os.path import isfile, join
import pandas as pd
from pandas import *
import random
import re
import sys 
import openpyxl
from openpyxl import load_workbook


class PivotError(Exception):
    """Base class for exceptions in this module."""
    pass 

def create_df(input):
    initial_df = pd.read_csv(input, sep=",", header=False, dtype='str')
    
    return initial_df
   
def gene_rollup_highest_impact(initial_df):
    sample_cols =  initial_df.filter(regex="RNA_FREQ")
    sample_col_array = sample_cols.columns
    col_array = []
    col_array.extend(sample_col_array)

    required_columns = set(["CHROM", "POS", "REF", "ANNOTATED_ALLELE", "GENE_SYMBOL", "HIGHEST_IMPACT"]+  col_array)
    for col in initial_df.columns:
        if col not in required_columns:
            del initial_df[col]
            
    melted_df = pd.melt(initial_df, id_vars=["CHROM", "POS", "REF", "ANNOTATED_ALLELE", "GENE_SYMBOL", "HIGHEST_IMPACT"], var_name="Sample", value_name="FREQ")
    melted_df = melted_df.fillna(".")

    filtered_df = melted_df[melted_df["FREQ"] != "."]

    pivoted_df = pd.pivot_table(filtered_df, rows=["GENE_SYMBOL", "Sample"], cols=["HIGHEST_IMPACT"], values=["FREQ"], aggfunc=np.count_nonzero, fill_value=0)

    pivoted_df = pivoted_df["FREQ"]

    pivoted_df["HIGH_initial_sum"]     = pivoted_df["HIGH"].map(int)
    pivoted_df["MODERATE_initial_sum"] = pivoted_df["MODERATE"].map(int) 
    pivoted_df["LOW_initial_sum"]      = pivoted_df["LOW"].map(int) 
    pivoted_df["MODIFIER_initial_sum"] = pivoted_df["MODIFIER"].map(int)
    
    # pivoted_df["Impact_complete"] = pivoted_df["HIGH"].map(str) + "|" + pivoted_df["MODERATE"].map(str) + "|" + pivoted_df["LOW"].map(str) + "|" + pivoted_df["MODIFIER"].map(str)
    pivoted_df["Impact_complete"] = pivoted_df["HIGH"].apply(lambda x: "h" * x) + pivoted_df["MODERATE"].apply(lambda x: "m" * x) + pivoted_df["LOW"].apply(lambda x: "l" * x) + pivoted_df["MODIFIER"].apply(lambda x: "x" * x)

    pivoted_df["Impact_score"] = pivoted_df["HIGH"] * 100000.0 + pivoted_df["MODERATE"] + pivoted_df["LOW"]/100000.0 + pivoted_df["MODIFIER"]/10**12

    del pivoted_df["HIGH"]
    del pivoted_df["MODERATE"]
    del pivoted_df["LOW"]
    del pivoted_df["MODIFIER"]
    
    expanded_df = pivoted_df.unstack()
    
    return expanded_df

def gene_rollup_damaging_impact(initial_df):   
    sample_cols =  initial_df.filter(regex="RNA_FREQ")
    sample_col_array = sample_cols.columns
    col_array = []
    col_array.extend(sample_col_array) 
    
    required_columns = set(["CHROM", "POS", "REF", "ANNOTATED_ALLELE", "GENE_SYMBOL", "Impact_Damaging"] +  col_array)
    for col in initial_df.columns:
        if col not in required_columns:
            del initial_df[col]
    
    melted_df = pd.melt(initial_df, id_vars=["CHROM", "POS", "REF", "ANNOTATED_ALLELE", "GENE_SYMBOL", "Impact_Damaging"], var_name="Sample", value_name="FREQ")
    filtered_df = melted_df.fillna(0)
    
    filtered_df = filtered_df[filtered_df["Impact_Damaging"] != "0"]
    filtered_df = filtered_df[filtered_df["FREQ"] != 0]

    filtered_df["Impact_Damaging"] = filtered_df["Impact_Damaging"].apply(lambda x: int(x))
    
    df_impact = pd.pivot_table(filtered_df, rows=["GENE_SYMBOL"], cols=["Sample"], values=["Impact_Damaging"], aggfunc=np.sum)

    return df_impact
    
def combine_dfs(df1, df2):
    df_combined = df1.join(df2, how='outer')
    df_combined = df_combined.fillna(0)
    
    ranked_df = calculate_ranks(df_combined)
    
    ranked_df = ranked_df.fillna("")
    
    return ranked_df
    
def calculate_ranks(df):
    df["Impact_complete_Rank"] = df["Impact_score"].sum(axis=1)
    df["Impact_Damaging_Rank"] = df["Impact_Damaging"].sum(axis=1)
    
    df["HIGH_sum"]     = df["HIGH_initial_sum"].sum(axis=1)
    df["MODERATE_sum"] = df["MODERATE_initial_sum"].sum(axis=1)
    df["LOW_sum"]      = df["LOW_initial_sum"].sum(axis=1)
    df["MODIFIER_sum"] = df["MODIFIER_initial_sum"].sum(axis=1)
    
    df["Impact_Damaging"] = df["Impact_Damaging"].applymap(lambda x: "" if x == 0.0 else x)
    
    df["Impact_Score"] = df["Impact_complete_Rank"]
    
    df = df.sort("Impact_complete_Rank", ascending=0)
    df = df.sort("Impact_Damaging_Rank", ascending=0)
    
    df["Impact_complete_Rank"] = df["Impact_complete_Rank"].rank(ascending=0, method="min")
    df["Impact_Damaging_Rank"] = df["Impact_Damaging_Rank"].rank(ascending=0, method="min")

    del df["Impact_score"]
    del df["HIGH_initial_sum"]
    del df["LOW_initial_sum"]
    del df["MODERATE_initial_sum"]
    del df["MODIFIER_initial_sum"]
    
    return df   
 
def rearrange_columns(output):
    ##change tuples to strings:
    lst = []
    for i in list(output.columns.values):
        if type(i) is tuple:
            i = "_".join(i)
        lst.append(i)
    output.columns = lst 
    
    ##change order of columns:
    lst = change_order(lst)    
    output = output.ix[:,lst]

    return output
    
def change_order(lst):
    gene_lst = []
    sum_lst = []
    level_lst = []
    impact_lst = []
    score_lst = []
    
    for i, header in enumerate(lst):    
        if re.search("GENE_SYMBOL", header):
            gene_lst.append(header)
        elif re.search("Impact_complete", header):
            impact_lst.append(header)
        elif re.search("sum", header):
            sum_lst.append(header)
        elif re.search("Score", header):
            score_lst.append(header)
        elif re.search("Impact_Damaging", header):
            level_lst.append(header)
   
    lst = gene_lst + impact_lst + sum_lst + score_lst + level_lst
    
    return lst

def style_workbook(output):
    wb = load_workbook(output)
    ws = wb.active
    
    for col in ws.columns:
        if re.search("Rank", col[0].value):
            format_cell_range(ws, col, '3366FF', 'FFFFFFFF')
        elif re.search("HIGH", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', 'FF0000')
        elif re.search("MODERATE", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', 'FFCC00')
        elif re.search("LOW", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', '00FF00')
        elif re.search("MODIFIER", col[0].value):
            format_cell_range(ws, col, 'FFFFFFFF', 'FF6600')

    wb.save(output)
    
def format_cell_range(ws, column, min_color, max_color):
    first_cell = str(column[1]).strip("(<>)").split(".")[-1] 
    last_cell = str(column[-1]).strip("(<>)").split(".")[-1] 
    cell = first_cell + ":" + last_cell
    
    ws.conditional_formatting.add2ColorScale(cell, 'min', None, min_color, 'max', None, max_color)
    
def process_files(input):
    initial_df = create_df(input)
    second_df = create_df(input)

    df1 = gene_rollup_highest_impact(initial_df)
    df2 = gene_rollup_damaging_impact(second_df)
    
    rollup_df = combine_dfs(df1, df2)
    
    rearranged_df = rearrange_columns(rollup_df)
    
    print "writing to excel file: {0}".format(output)
    writer = ExcelWriter(output)
    
    rearranged_df.to_excel(writer, "Variant_output", index=True, merge_cells = 0)  

    print "saving file"
    writer.save() 
    
    print "styling workbook"
    style_workbook(output)

    print "done"
    
    
if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pd.set_option('chained_assignment', None)
     
    if len(sys.argv) != 3:
        print "Invalid arguments."
        print "Usage: pivot.py [in_csv_file] [out_excel_file]"
        print "Example: pivot.py ../Rhim_input.csv ../Rhim_output.xlsx"
        exit(1)
    else:
        input    = script_dir + "/" + sys.argv[1]
        output   = script_dir + "/" + sys.argv[2]
 

    if not os.path.isfile(input):
        print "Error. Specified input file {0} does not exist".format(input)
        exit(1)

    process_files(input)